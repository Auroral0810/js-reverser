import json
import execjs
import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import execjs

cookies = {
    'Hm_lvt_6088e7f72f5a363447d4bafe03026db8': '1744543815',
    'HMACCOUNT': '764A7B05229BE584',
    'Hm_lpvt_6088e7f72f5a363447d4bafe03026db8': '1744545118',
}

headers = {
    'Accept': '*/*',
    'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
    'Cache-Control': 'no-cache',
    'Connection': 'keep-alive',
    'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
    'Origin': 'https://www.aqistudy.cn',
    'Pragma': 'no-cache',
    'Referer': 'https://www.aqistudy.cn/historydata/monthdata.php?city=%E4%B8%8A%E6%B5%B7',
    'Sec-Fetch-Dest': 'empty',
    'Sec-Fetch-Mode': 'cors',
    'Sec-Fetch-Site': 'same-origin',
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/135.0.0.0 Safari/537.36',
    'X-Requested-With': 'XMLHttpRequest',
    'sec-ch-ua': '"Google Chrome";v="135", "Not-A.Brand";v="8", "Chromium";v="135"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': '"macOS"',
}


city = "哈尔滨"
month = "202503"
city_param = {
    "city": city,
    "month": month
}


with open('实现.js', 'r', encoding='utf-8') as f:
    js = f.read()
ctx = execjs.compile(js)

hA4Nse2cT = ctx.call('poPBVxzNuafY8Yu',"GETDAYDATA", city_param)
print(hA4Nse2cT)
# 构造请求参数
data = {
    'hA4Nse2cT': hA4Nse2cT
}# 发送请求获取响应
response = requests.post('https://www.aqistudy.cn/historydata/api/historyapi.php', cookies=cookies, headers=headers, data=data).text
# print(response)
result = ctx.call('dxvERkeEvHbS', response)

# 修复：将字符串解析为JSON对象
result_dict = json.loads(result)

# 解析数据
data = result_dict["result"]['data']['items']
print(data)

# 创建一个Excel写入器
excel_writer = pd.ExcelWriter(f'{city}空气质量数据_{month}.xlsx', engine='openpyxl')

# 处理每日详细数据
daily_data = []
for item in data:
    daily_data.append({
        '日期': item['time_point'],
        'AQI': item['aqi'],
        'PM2.5': item['pm2_5'],
        'PM10': item['pm10'],
        'SO2': item['so2'],
        'NO2': item['no2'],
        'CO': item['co'],
        'O3': item['o3'],
        '排名': item['rank'],
        '空气质量': item['quality']
    })

# 创建DataFrame并写入Excel
daily_df = pd.DataFrame(daily_data)
daily_df.to_excel(excel_writer, sheet_name='每日空气质量数据', index=False)

# 计算统计数据
aqi_values = [item['aqi'] for item in data]
pm25_values = [item['pm2_5'] for item in data]
pm10_values = [item['pm10'] for item in data]

# 创建统计数据表
stats_data = {
    '指标': ['平均AQI', '最大AQI', '最小AQI', '平均PM2.5', '平均PM10', '数据点数量'],
    '值': [
        sum(aqi_values) / len(aqi_values) if aqi_values else 0,
        max(aqi_values) if aqi_values else 0,
        min(aqi_values) if aqi_values else 0,
        sum(pm25_values) / len(pm25_values) if pm25_values else 0,
        sum(pm10_values) / len(pm10_values) if pm10_values else 0,
        len(data)
    ]
}
stats_df = pd.DataFrame(stats_data)
stats_df.to_excel(excel_writer, sheet_name='数据统计', index=False)

# 统计空气质量级别分布
quality_counts = {}
for item in data:
    quality = item['quality']
    if quality in quality_counts:
        quality_counts[quality] += 1
    else:
        quality_counts[quality] = 1

quality_data = {
    '空气质量级别': list(quality_counts.keys()),
    '天数': list(quality_counts.values())
}
quality_df = pd.DataFrame(quality_data)
quality_df.to_excel(excel_writer, sheet_name='空气质量分布', index=False)

# 保存Excel文件
excel_writer.close()

# 美化Excel文件
from openpyxl import load_workbook
wb = load_workbook(f'{city}空气质量数据_{month}.xlsx')

# 设置每个工作表的样式
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    # 设置标题行样式
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

# 保存美化后的Excel
wb.save(f'{city}空气质量数据_{month}.xlsx')

print(f"数据已成功导出到'{city}空气质量数据_{month}.xlsx'文件中")
