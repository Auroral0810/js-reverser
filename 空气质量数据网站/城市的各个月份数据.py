import json
import execjs
import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import execjs

cookies = {
    'Hm_lvt_6088e7f72f5a363447d4bafe03026db8': '1744543815',
    'HMACCOUNT': '764A7B05229BE584',
    'Hm_lpvt_6088e7f72f5a363447d4bafe03026db8': '1744545118',
}

headers = {
    'Accept': '*/*',
    'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
    'Cache-Control': 'no-cache',
    'Connection': 'keep-alive',
    'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
    'Origin': 'https://www.aqistudy.cn',
    'Pragma': 'no-cache',
    'Referer': 'https://www.aqistudy.cn/historydata/monthdata.php?city=%E4%B8%8A%E6%B5%B7',
    'Sec-Fetch-Dest': 'empty',
    'Sec-Fetch-Mode': 'cors',
    'Sec-Fetch-Site': 'same-origin',
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/135.0.0.0 Safari/537.36',
    'X-Requested-With': 'XMLHttpRequest',
    'sec-ch-ua': '"Google Chrome";v="135", "Not-A.Brand";v="8", "Chromium";v="135"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': '"macOS"',
}


city = "哈尔滨"
# month = "202503"
city_param = {
    "city": city,
    # "month": month
}


with open('实现.js', 'r', encoding='utf-8') as f:
    js = f.read()
ctx = execjs.compile(js)

hA4Nse2cT = ctx.call('poPBVxzNuafY8Yu',"GETMONTHDATA", city_param)
print(hA4Nse2cT)
# 构造请求参数
data = {
    'hA4Nse2cT': hA4Nse2cT
}# 发送请求获取响应
response = requests.post('https://www.aqistudy.cn/historydata/api/historyapi.php', cookies=cookies, headers=headers, data=data).text
print(response)
result = ctx.call('dxvERkeEvHbS', response)

# 修复：将字符串解析为JSON对象
result_dict = json.loads(result)

# 假设我们已经有了JSON数据，并将其保存在变量result_dict中
# 首先解析数据
data = result_dict["result"]['data']

# 创建一个Excel写入器
excel_writer = pd.ExcelWriter('北京空气质量数据.xlsx', engine='openpyxl')

# 1. 处理月度详细数据
items_df = pd.DataFrame(data['items'])
items_df.to_excel(excel_writer, sheet_name='月度详细数据', index=False)

# 2. 处理污染级别统计
level_data = data['level']
level_df = pd.DataFrame({
    '污染级别': ['优(Level 1)', '良(Level 2)', '轻度污染(Level 3)', '中度污染(Level 4)', '重度污染(Level 5)', '严重污染(Level 6)'],
    '天数': [level_data['level1'], level_data['level2'], level_data['level3'],
            level_data['level4'], level_data['level5'], level_data['level6']]
})
level_df.to_excel(excel_writer, sheet_name='污染级别统计', index=False)

# 3. 处理AQI时间序列数据
aqi_df = pd.DataFrame(data['datas'])
aqi_df['date'] = pd.to_datetime(aqi_df['x'], unit='ms')
aqi_df = aqi_df[['date', 'y', 'level']]
aqi_df.columns = ['日期', 'AQI值', '污染级别']
aqi_df.to_excel(excel_writer, sheet_name='AQI时间序列', index=False)

# 4. 处理最小和最大AQI数据
min_aqi_df = pd.DataFrame(data['datas_min'])
min_aqi_df['date'] = pd.to_datetime(min_aqi_df['x'], unit='ms')
min_aqi_df = min_aqi_df[['date', 'y', 'level']]
min_aqi_df.columns = ['日期', '最小AQI值', '污染级别']

max_aqi_df = pd.DataFrame(data['datas_max'])
max_aqi_df['date'] = pd.to_datetime(max_aqi_df['x'], unit='ms')
max_aqi_df = max_aqi_df[['date', 'y', 'level']]
max_aqi_df.columns = ['日期', '最大AQI值', '污染级别']

# 合并最小和最大AQI数据
aqi_range_df = pd.merge(min_aqi_df, max_aqi_df, on='日期', suffixes=('_最小', '_最大'))
aqi_range_df.to_excel(excel_writer, sheet_name='AQI范围', index=False)

# 5. 提取摘要统计数据
summary_data = {
    '指标': ['平均AQI', '最大AQI', '最小AQI', '数据点数量'],
    '值': [data['avg'], data['max'], data['min'], data['num']]
}
summary_df = pd.DataFrame(summary_data)
summary_df.to_excel(excel_writer, sheet_name='数据摘要', index=False)

# 6. 处理不同污染级别的数据分布
levels_df = pd.DataFrame()
for i in range(1, 7):
    df = pd.DataFrame(data[f'datas{i}'])
    if not df.empty:
        df['date'] = pd.to_datetime(df['x'], unit='ms')
        df = df[['date', 'y']]
        df.columns = ['日期', f'Level {i}天数']

        if levels_df.empty:
            levels_df = df
        else:
            levels_df = pd.merge(levels_df, df, on='日期', how='outer')

levels_df.to_excel(excel_writer, sheet_name='污染级别分布', index=False)

# 保存Excel文件
excel_writer.close()

# 美化Excel文件
from openpyxl import load_workbook
wb = load_workbook('北京空气质量数据.xlsx')

# 设置每个工作表的样式
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    # 设置标题行样式
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

# 保存美化后的Excel
wb.save(f'{city}空气质量数据.xlsx')

print(f"数据已成功导出到'{city}空气质量数据.xlsx'文件中")
